from flask import Flask, render_template, redirect, request
from flask_sqlalchemy import SQLAlchemy
import openpyxl

app = Flask(__name__)
app.config['SECRET_KEY'] = "Super Secret Key"
app.config['SQLALCHEMY_DATABASE_URI'] = "mysql://root:@localhost/colleges"
db = SQLAlchemy(app)

path = "D:\Map\Book1.xlsx"
wb_obj = openpyxl.open(path)
sheet_obj = wb_obj.get_sheet_by_name("Sheet1")

class Colleges(db.Model):
    __tablename__ = "colleges"
    id = db.Column('sl_', db.Integer, primary_key=True)
    institute = db.Column('institute', db.String(255))
    district = db.Column('district', db.String(255))
    subdivision=db.Column('subdivision',db.String(255))
    block=db.Column('block',db.String(255))
    status=db.Column('status',db.String(255))
    address=db.Column('address',db.String(255))
    contact=db.Column('contact',db.String(255))
    email=db.Column('email',db.String(255))
    lat = db.Column('lat',db.Float(20))
    lng = db.Column('lng',db.Float(20))

allcolleges = Colleges.query.all();
print(allcolleges);
for i in range(1, 41):
    id = sheet_obj.cell(row=i, column=1).value
    institute = sheet_obj.cell(row=i, column=2).value
    district = sheet_obj.cell(row=i, column=3).value
    sub_div = sheet_obj.cell(row=i, column=4).value
    block = sheet_obj.cell(row=i, column=5).value
    status = sheet_obj.cell(row=i, column=6).value
    address = sheet_obj.cell(row=i, column=7).value
    contact = sheet_obj.cell(row=i, column=8).value
    email = sheet_obj.cell(row=i, column=9).value
    lat = sheet_obj.cell(row=i, column=10).value
    lng = sheet_obj.cell(row=i, column=11).value
    adddata = Colleges(id=id, institute=institute, district=district, subdivision=sub_div,block=block,status=status, address=address, contact=contact, email=email, lat=lat, lng=lng);
    db.session.add(adddata);
    db.session.commit();
    print(block)


